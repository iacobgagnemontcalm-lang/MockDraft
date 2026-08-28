#!/usr/bin/env python3
"""Daily rankings refresh for DinoMock.

Regenerates the bundled NFL ranking CSVs from their live sources:

  rankings.csv                    FantasyPros PPR consensus cheatsheet (ECR + tiers)
  adp_rankings.csv                consensus PPR ADP, averaged across the
                                  FFC / ESPN / Yahoo public feeds
  {sleeper,espn,yahoo}_rankings.csv  one file per platform: that platform's own
                                  live daily ADP (Sleeper / ESPN / Yahoo public
                                  APIs) for the draft order, plus the Landmine
                                  score from the "Abusing Fantasy Draft
                                  Rankings" sheet, which is the only place that
                                  hand-authored score exists

Each source is fetched independently — if one fails, its CSV is left
untouched (stale but valid) and the others still update. The script exits
non-zero only when every source fails.

CFL CSVs (cfl_rankings.csv / cfl_adp_rankings.csv) have no automated
source and are maintained by hand.
"""

import csv
import io
import json
import re
import sys
import unicodedata

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

REPO_ROOT = re.sub(r"/scripts$", "", __import__("os").path.dirname(__import__("os").path.abspath(__file__))) or "."

UA = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36"
}

ECR_URL = "https://www.fantasypros.com/nfl/rankings/ppr-cheatsheets.php"
SHEET_ID = "1HTixsrRtIIpnUafVkOIhET83vCFjKXSUGiG24-5jTHY"
SHEET_XLSX_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"

# Refuse to overwrite a CSV with a suspiciously small player pool — a layout
# change or half-rendered page should never clobber good data.
MIN_ECR_PLAYERS = 300
MIN_ADP_PLAYERS = 100
MIN_SHEET_PLAYERS = 100


def write_csv(path, header, rows):
    with open(path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerows(rows)


def update_ecr():
    """rankings.csv from the FantasyPros PPR cheatsheet's embedded ecrData."""
    html = requests.get(ECR_URL, headers=UA, timeout=60).text
    m = re.search(r"var\s+ecrData\s*=\s*(\{.*?\})\s*;", html, re.S)
    if not m:
        raise RuntimeError("ecrData blob not found on cheatsheet page")
    players = json.loads(m.group(1))["players"]
    if len(players) < MIN_ECR_PLAYERS:
        raise RuntimeError(f"only {len(players)} players in ecrData — refusing to overwrite")

    players.sort(key=lambda p: p["rank_ecr"])
    rows = []
    for p in players:
        rows.append([
            p["rank_ecr"],
            p.get("tier", ""),
            p["player_name"],
            p.get("player_team_id", ""),
            p.get("pos_rank", p.get("player_position_id", "")),
            p.get("player_bye_week", ""),
            p.get("rank_min", ""),
            p.get("rank_max", ""),
            p.get("rank_ave", ""),
            p.get("rank_std", ""),
            p.get("player_ecr_delta", ""),
        ])
    write_csv(
        f"{REPO_ROOT}/rankings.csv",
        ["RK", "TIERS", "PLAYER NAME", "TEAM", "POS", "BYE WEEK",
         "BEST", "WORST", "AVG.", "STD.DEV", "ECR VS. ADP"],
        rows,
    )
    return len(rows)


# FFC names defenses by city ("Seattle Defense"); rankings.csv uses full team
# names ("Seattle Seahawks"). index.html matches ADP rows to players by exact
# name, so translate.
FFC_DST_NAMES = {
    "Arizona": "Arizona Cardinals", "Atlanta": "Atlanta Falcons",
    "Baltimore": "Baltimore Ravens", "Buffalo": "Buffalo Bills",
    "Carolina": "Carolina Panthers", "Chicago": "Chicago Bears",
    "Cincinnati": "Cincinnati Bengals", "Cleveland": "Cleveland Browns",
    "Dallas": "Dallas Cowboys", "Denver": "Denver Broncos",
    "Detroit": "Detroit Lions", "Green Bay": "Green Bay Packers",
    "Houston": "Houston Texans", "Indianapolis": "Indianapolis Colts",
    "Jacksonville": "Jacksonville Jaguars", "Kansas City": "Kansas City Chiefs",
    "LA Chargers": "Los Angeles Chargers", "LA Rams": "Los Angeles Rams",
    "Las Vegas": "Las Vegas Raiders", "Miami": "Miami Dolphins",
    "Minnesota": "Minnesota Vikings", "New England": "New England Patriots",
    "New Orleans": "New Orleans Saints", "NY Giants": "New York Giants",
    "NY Jets": "New York Jets", "Philadelphia": "Philadelphia Eagles",
    "Pittsburgh": "Pittsburgh Steelers", "San Francisco": "San Francisco 49ers",
    "Seattle": "Seattle Seahawks", "Tampa Bay": "Tampa Bay Buccaneers",
    "Tennessee": "Tennessee Titans", "Washington": "Washington Commanders",
}


def adp_from_ffc():
    """ADP rows from Fantasy Football Calculator's public JSON API."""
    url = f"https://fantasyfootballcalculator.com/api/v1/adp/ppr?teams=12&year={SEASON}"
    data = requests.get(url, headers=UA, timeout=60).json()
    rows = []
    for p in data.get("players") or []:
        name = (p.get("name") or "").strip()
        if not name:
            continue
        m = re.fullmatch(r"(.+?)\s+Defense", name)
        if m:
            name = FFC_DST_NAMES.get(m.group(1), name)
        try:
            adp = float(p.get("adp"))
        except (TypeError, ValueError):
            continue
        if adp <= 0:
            continue
        rows.append({"name": name, "team": p.get("team") or "",
                     "pos": p.get("position") or "", "adp": adp})
    return rows, "FantasyFootballCalculator"


def _norm_name(n):
    """Suffix- and accent-insensitive name key (mirrors index.html's fuzzyGet)."""
    n = unicodedata.normalize("NFKD", n).encode("ascii", "ignore").decode()
    n = n.lower().replace(".", "").replace("'", "")
    n = re.sub(r"\s+(jr|sr|ii|iii|iv|v)$", "", n.strip())
    return re.sub(r"\s+", " ", n)


# ── Platform draft order ───────────────────────────────────────────────────
#
# {sleeper,espn,yahoo}_rankings.csv carry two different things:
#
#   SITE RANK  the order that platform's rooms actually draft in — this moves
#              every day, so it is pulled live from each platform's own public
#              API below.
#   LANDMINE   a hand-authored 0-10 reach-risk score with no API anywhere; it
#              only exists in the community "Abusing Draft Rankings" sheet,
#              which its author republishes about once a week.
#
# So the sheet is still read (for LANDMINE, and as the fallback SITE RANK when
# a platform's API is unreachable), but it no longer gates how often the draft
# order refreshes.

# Sheet tab -> output CSV basename.
SITE_TABS = {
    "Sleeper PPR": "sleeper",
    "ESPN PPR": "espn",
    "Yahoo PPR": "yahoo",
}

# NFL seasons span the new year, so Jan/Feb still belong to the previous one.
SEASON = (lambda d: d.year if d.month >= 3 else d.year - 1)(
    __import__("datetime").date.today()
)

# A live feed with a tiny player pool means the endpoint changed shape — keep
# the sheet's ranks rather than publishing a half-empty draft order.
MIN_LIVE_ADP = 150

# How many players to publish per platform. Sleeper ranks every rostered
# player and ESPN serves whatever limit it is given; past a few hundred the
# rows are undrafted noise.
SITE_POOL = 400

ESPN_POS = {1: "QB", 2: "RB", 3: "WR", 4: "TE", 5: "K", 16: "DST"}

# Sleeper lists defenses under a team abbreviation with no full_name, but
# rankings.csv names them in full and index.html matches by exact name.
DST_BY_ABBR = {
    "ARI": "Arizona Cardinals", "ATL": "Atlanta Falcons", "BAL": "Baltimore Ravens",
    "BUF": "Buffalo Bills", "CAR": "Carolina Panthers", "CHI": "Chicago Bears",
    "CIN": "Cincinnati Bengals", "CLE": "Cleveland Browns", "DAL": "Dallas Cowboys",
    "DEN": "Denver Broncos", "DET": "Detroit Lions", "GB": "Green Bay Packers",
    "HOU": "Houston Texans", "IND": "Indianapolis Colts", "JAX": "Jacksonville Jaguars",
    "KC": "Kansas City Chiefs", "LAC": "Los Angeles Chargers", "LAR": "Los Angeles Rams",
    "LV": "Las Vegas Raiders", "MIA": "Miami Dolphins", "MIN": "Minnesota Vikings",
    "NE": "New England Patriots", "NO": "New Orleans Saints", "NYG": "New York Giants",
    "NYJ": "New York Jets", "PHI": "Philadelphia Eagles", "PIT": "Pittsburgh Steelers",
    "SEA": "Seattle Seahawks", "SF": "San Francisco 49ers", "TB": "Tampa Bay Buccaneers",
    "TEN": "Tennessee Titans", "WAS": "Washington Commanders", "WSH": "Washington Commanders",
}
ESPN_TEAMS = {
    0: "FA", 1: "ATL", 2: "BUF", 3: "CHI", 4: "CIN", 5: "CLE", 6: "DAL",
    7: "DEN", 8: "DET", 9: "GB", 10: "TEN", 11: "IND", 12: "KC", 13: "LV",
    14: "LAR", 15: "MIA", 16: "MIN", 17: "NE", 18: "NO", 19: "NYG", 20: "NYJ",
    21: "PHI", 22: "ARI", 23: "PIT", 24: "LAC", 25: "SF", 26: "SEA", 27: "TB",
    28: "WSH", 29: "CAR", 30: "JAX", 33: "BAL", 34: "HOU",
}


def sleeper_adp():
    """Sleeper's live draft order, from the public players endpoint.

    Sleeper exposes no ADP anywhere public — their GraphQL schema has no
    `adp_data` field — so this uses `search_rank`, the order Sleeper's own
    draft board lists players in. It is the same thing SITE RANK means for the
    other two platforms, and it moves as Sleeper re-rates players.
    """
    players = requests.get(
        "https://api.sleeper.app/v1/players/nfl", headers=UA, timeout=120
    ).json()

    out = []
    for p in players.values():
        rank = p.get("search_rank")
        if not isinstance(rank, (int, float)) or rank >= 9999:
            continue
        name = (p.get("full_name") or "").strip()
        if not name and p.get("position") == "DEF":
            abbr = (p.get("team") or "").strip().upper()
            name = DST_BY_ABBR.get(abbr, abbr)
        if not name:
            continue
        out.append({
            "name": name,
            "team": p.get("team") or "",
            "pos": p.get("position") or "",
            "adp": float(rank),
        })
    return out, "Sleeper search_rank"


def espn_adp():
    """ESPN's live PPR ADP from the public default-league player endpoint."""
    url = (f"https://lm-api-reads.fantasy.espn.com/apis/v3/games/ffl/seasons/"
           f"{SEASON}/segments/0/leaguedefaults/3")
    resp = requests.get(
        url,
        params={"view": "kona_player_info"},
        headers={
            **UA,
            "Accept": "application/json",
            "x-fantasy-filter": json.dumps({
                "players": {
                    "limit": SITE_POOL,
                    "sortDraftRanks": {
                        "sortPriority": 100, "sortAsc": True, "value": "PPR",
                    },
                }
            }),
        },
        timeout=120,
    )
    resp.raise_for_status()
    out = []
    for entry in resp.json().get("players") or []:
        p = entry.get("player") or {}
        name = (p.get("fullName") or "").strip()
        if not name:
            continue
        adp = (p.get("ownership") or {}).get("averageDraftPosition")
        if not isinstance(adp, (int, float)) or adp <= 0:
            # Undrafted in ESPN rooms yet — fall back to their static PPR rank.
            adp = ((p.get("draftRanksByRankType") or {}).get("PPR") or {}).get("rank")
        if not isinstance(adp, (int, float)) or adp <= 0:
            continue
        pos = ESPN_POS.get(p.get("defaultPositionId"), "")
        out.append({
            "name": name,
            "team": ESPN_TEAMS.get(p.get("proTeamId"), ""),
            "pos": pos,
            "adp": float(adp),
        })
    return out, "ESPN averageDraftPosition"


def _yahoo_consider(p, out):
    """Record a Yahoo player dict if it carries both a name and an ADP."""
    name = p.get("name")
    full = name.get("full") if isinstance(name, dict) else None
    da = p.get("draft_analysis")
    if isinstance(da, list):
        da = next((d for d in da if isinstance(d, dict) and "average_pick" in d), None)
    if not full or not isinstance(da, dict):
        return
    try:
        adp = float(da.get("average_pick"))
    except (TypeError, ValueError):
        return
    if adp <= 0:
        return
    full = full.strip()
    out[full] = {
        "name": full,
        "team": (p.get("editorial_team_abbr") or "").upper(),
        "pos": p.get("display_position") or "",
        "adp": adp,
    }


def _merge_fragments(lst):
    """Merge the dict fragments nested in `lst` into one view.

    Yahoo's v2 JSON splits a single player across sibling dicts inside nested
    lists (``player: [[{name}, {team}, ...], {draft_analysis}]``). Returns the
    merged dict and how many fragments carried a name — more than one means
    this list holds several players, not one player's fragments.
    """
    merged, names, stack = {}, 0, list(lst)
    while stack:
        cur = stack.pop()
        if isinstance(cur, list):
            stack.extend(cur)
        elif isinstance(cur, dict):
            if "name" in cur:
                names += 1
            for k, v in cur.items():
                merged.setdefault(k, v)
    return merged, names


def _walk_yahoo(node, out):
    """Collect {name, team, pos, adp} from Yahoo's variably-nested JSON.

    The wrapping differs by endpoint and format — flat player dicts under
    ``format=json_f``, fragment lists otherwise — so recurse and handle both
    rather than indexing into a fixed shape.
    """
    if isinstance(node, dict):
        _yahoo_consider(node, out)
        for v in node.values():
            _walk_yahoo(v, out)
    elif isinstance(node, list):
        merged, names = _merge_fragments(node)
        if names == 1:
            _yahoo_consider(merged, out)
        for v in node:
            _walk_yahoo(v, out)


def yahoo_adp():
    """Yahoo's live ADP from the read-only public API their draft app uses."""
    base = "https://pub-api-ro.fantasysports.yahoo.com/fantasy/v2/game/nfl/players"
    found, barren = {}, 0
    for start in range(0, 3 * SITE_POOL, 25):
        path = (f"{base};position=ALL;start={start};count=25;sort=rank_season;"
                f"search=;out=draft_analysis;ranks=season/draft_analysis")
        resp = requests.get(path, params={"format": "json_f"},
                            headers={**UA, "Accept": "application/json"}, timeout=60)
        if resp.status_code == 404:
            break
        resp.raise_for_status()
        body = resp.text
        before = len(found)
        _walk_yahoo(json.loads(body), found)
        if '"full"' not in body:
            break  # no players at all — end of the list
        # Yahoo only publishes a numeric average_pick for the players actually
        # being drafted (~225); past that every row reads "-". A single barren
        # page is normal mid-list, a run of them means the ADP has run out.
        barren = barren + 1 if len(found) == before else 0
        if barren >= 3 or len(found) >= SITE_POOL:
            break
    return list(found.values()), "Yahoo average_pick"


LIVE_ADP = {"sleeper": sleeper_adp, "espn": espn_adp, "yahoo": yahoo_adp}


def _rankings_index():
    """{norm name: (exact name, TEAM, BYE, POS)} from rankings.csv."""
    idx = {}
    with open(f"{REPO_ROOT}/rankings.csv") as f:
        for row in csv.DictReader(f):
            idx.setdefault(_norm_name(row["PLAYER NAME"]), (
                row["PLAYER NAME"], row["TEAM"],
                row["BYE WEEK"], re.sub(r"\d+$", "", row["POS"]),
            ))
    return idx


def _live_name(rec):
    """Canonical name for a live feed row — defenses resolve via their team.

    Each platform names a defense differently ("Texans", "Rams", "HOU");
    rankings.csv and index.html use the full team name.
    """
    if (rec.get("pos") or "").upper() in ("DST", "DEF", "D/ST"):
        full = DST_BY_ABBR.get((rec.get("team") or "").upper())
        if full:
            return full
    return rec["name"]


def update_adp():
    """adp_rankings.csv — consensus ADP averaged over every live source.

    FantasyPros used to serve this table server-side, but the ADP page is now
    rendered client-side (the only table left in the HTML is the "experts"
    credits table, which is what the old scraper was silently picking up) and
    their JSON API rejects unkeyed requests. Their cheatsheet `ecrData` blob
    carries expert rank and ownership but no ADP either.

    So the consensus is built here instead, by averaging each player's pick
    across the public feeds that do work: FantasyFootballCalculator's mock
    drafts plus ESPN's and Yahoo's real-draft ADP. A player is included on
    the strength of any one source; where several have him they are averaged,
    which is what makes it a consensus rather than a single room's habits.
    """
    idx = _rankings_index()
    picks, seen = {}, {}
    used, failed = [], []

    for source, fn in (("FFC", adp_from_ffc), ("ESPN", espn_adp), ("Yahoo", yahoo_adp)):
        try:
            rows, label = fn()
            if len(rows) < MIN_ADP_PLAYERS:
                raise RuntimeError(f"only {len(rows)} rows")
        except Exception as e:
            failed.append(f"{source} ({e})")
            continue
        for r in rows:
            key = _norm_name(_live_name(r))
            picks.setdefault(key, []).append(r["adp"])
            seen.setdefault(key, r)
        used.append(f"{label} ({len(rows)})")

    if not picks:
        raise RuntimeError(f"every ADP source failed: {'; '.join(failed)}")

    blended = []
    for key, vals in picks.items():
        exact, team, bye, pos = idx.get(key, (None, "", "", ""))
        rec = seen[key]
        name = exact or _live_name(rec)
        # Reproduce the bundled "Name   TEAM (BYE)" suffix, which
        # index.html's loadADPCSVText strips when matching names.
        team = team or (rec.get("team") or "").upper()
        suffix = f"   {team}" + (f" ({bye})" if bye else "") if team else ""
        blended.append((sum(vals) / len(vals), name + suffix, pos or rec.get("pos") or ""))

    blended.sort(key=lambda b: b[0])
    if len(blended) < MIN_ADP_PLAYERS:
        raise RuntimeError(f"only {len(blended)} blended ADP rows — refusing to overwrite")

    write_csv(
        f"{REPO_ROOT}/adp_rankings.csv",
        ["Rank", "Player (Bye)", "POS", "AVG"],
        [[i, name, pos, round(adp, 1)] for i, (adp, name, pos) in enumerate(blended, 1)],
    )
    print(f"     ADP sources: {', '.join(used) or 'none'}")
    if failed:
        print(f"     ADP sources unavailable: {'; '.join(failed)}", file=sys.stderr)
    return len(blended)




def read_sheet(canon):
    """{platform: [row dicts]} from the weekly Abusing Draft Rankings sheet."""
    resp = requests.get(SHEET_XLSX_URL, headers=UA, timeout=120)
    resp.raise_for_status()
    wb = load_workbook(io.BytesIO(resp.content), read_only=True, data_only=True)

    def iv(x):
        try:
            return int(float(x))
        except (TypeError, ValueError):
            return ""

    tables = {}
    for tab_pattern, out_name in SITE_TABS.items():
        tab = next((n for n in wb.sheetnames if n.strip().lower() == tab_pattern.lower()), None)
        if tab is None:
            raise RuntimeError(f"no '{tab_pattern}' tab — sheets: {wb.sheetnames}")
        all_rows = list(wb[tab].iter_rows(values_only=True))
        header = [str(c).strip().lower() if c is not None else "" for c in all_rows[0]]

        def col(name):
            if name not in header:
                raise RuntimeError(f"column '{name}' not found in '{tab}' header: {header}")
            return header.index(name)

        i_name, i_team, i_bye, i_pos = col("name"), col("team"), col("bye"), col("pos")
        i_fp = col("fantasypros")
        i_site = i_fp + 1  # the platform's rank column always follows FantasyPros
        i_lm = col("landmine")
        width = max(i_lm, i_site) + 1

        rows = []
        for r in all_rows[1:]:
            r = list(r) + [None] * max(0, width - len(r))
            if not r[i_name] or r[i_fp] is None or r[i_site] is None:
                continue
            name = str(r[i_name]).strip()
            rows.append({
                "name": canon.get(_norm_name(name), name),
                "team": r[i_team] or "",
                "pos": r[i_pos] or "",
                "bye": iv(r[i_bye]),
                "fp": iv(r[i_fp]),
                "site": iv(r[i_site]),
                "lm": r[i_lm] if r[i_lm] is not None else "",
            })

        if len(rows) < MIN_SHEET_PLAYERS:
            raise RuntimeError(f"only {len(rows)} rows parsed from '{tab}' — refusing to overwrite")
        tables[out_name] = rows
    return tables


def write_site_csv(out_name, rows):
    write_csv(
        f"{REPO_ROOT}/{out_name}_rankings.csv",
        ["PLAYER NAME", "TEAM", "POS", "BYE", "FP RANK", "SITE RANK", "LANDMINE", "ADP"],
        [[r["name"], r["team"], r["pos"], r["bye"], r["fp"], r["site"], r["lm"],
          r.get("adp", "")] for r in rows],
    )


def merge_live(sheet_rows, live_rows, canon):
    """Re-rank the platform's players by today's live ADP.

    Ordering comes entirely from the live feed, so SITE RANK is a dense 1..N
    draft order just like the sheet's was. LANDMINE and FP RANK are carried
    across by name from the sheet, which is the only place they exist.
    """
    by_name = {_norm_name(r["name"]): r for r in sheet_rows}
    live_rows = sorted(live_rows, key=lambda r: r["adp"])[:SITE_POOL]

    merged = []
    for i, live in enumerate(live_rows, 1):
        key = _norm_name(_live_name(live))
        sheet = by_name.get(key, {})
        merged.append({
            "name": canon.get(key, sheet.get("name") or _live_name(live)),
            "team": sheet.get("team") or live["team"],
            "pos": sheet.get("pos") or live["pos"],
            "bye": sheet.get("bye", ""),
            "fp": sheet.get("fp", ""),
            "site": i,
            "lm": sheet.get("lm", ""),
            "adp": round(live["adp"], 1),
        })
    return merged


def update_site_ranks():
    """{sleeper,espn,yahoo}_rankings.csv — live daily ADP + the sheet's Landmine."""
    # Canonical spellings from rankings.csv — the app matches by exact name
    canon = {}
    with open(f"{REPO_ROOT}/rankings.csv") as f:
        for row in csv.DictReader(f):
            canon.setdefault(_norm_name(row["PLAYER NAME"]), row["PLAYER NAME"])

    try:
        tables = read_sheet(canon)
    except Exception as e:
        print(f"     note: Landmine sheet unavailable ({e}) — live ranks only",
              file=sys.stderr)
        tables = {name: [] for name in SITE_TABS.values()}

    total, failures = 0, []
    for out_name in SITE_TABS.values():
        sheet_rows = tables.get(out_name) or []
        try:
            live_rows, source = LIVE_ADP[out_name]()
            if len(live_rows) < MIN_LIVE_ADP:
                raise RuntimeError(f"only {len(live_rows)} players from {source}")
            rows = merge_live(sheet_rows, live_rows, canon)
            print(f"     {out_name}_rankings.csv: {len(rows)} players (live — {source})")
        except Exception as e:
            if not sheet_rows:
                failures.append(f"{out_name} ({e})")
                print(f"     FAIL {out_name}: live ADP failed ({e}) and no sheet data",
                      file=sys.stderr)
                continue
            rows = sheet_rows
            print(f"     {out_name}_rankings.csv: {len(rows)} players "
                  f"(weekly sheet — live ADP failed: {e})", file=sys.stderr)
        write_site_csv(out_name, rows)
        total += len(rows)

    if failures:
        raise RuntimeError("; ".join(failures))
    return total


def main():
    results = {}
    for label, fn in [
        ("rankings.csv (FantasyPros ECR)", update_ecr),
        ("adp_rankings.csv (consensus ADP)", update_adp),
        ("site rank CSVs (live platform ADP)", update_site_ranks),
    ]:
        try:
            results[label] = fn()
            print(f"OK   {label}: {results[label]} players")
        except Exception as e:
            results[label] = None
            print(f"FAIL {label}: {e}", file=sys.stderr)

    if all(v is None for v in results.values()):
        print("All sources failed — nothing updated", file=sys.stderr)
        sys.exit(1)
    failed = [k for k, v in results.items() if v is None]
    if failed:
        print(f"WARNING: {len(failed)} source(s) failed, kept previous data: {', '.join(failed)}",
              file=sys.stderr)


if __name__ == "__main__":
    main()
